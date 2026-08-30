
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.formatting.rule import CellIsRule

st.set_page_config(
    page_title="Cost Sheet Pro",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =========================
# STANDARD COST LIBRARY
# =========================
COST_LIBRARY = {
    "Direct Material": [
        "Opening Raw Material Stock", "Raw Material Purchases",
        "Closing Raw Material Stock", "Raw Materials", "Basic Materials", "Components", "Sub-assemblies",
        "Purchased Parts", "Packing Material - Direct", "Consumables - Direct",
        "Stores - Direct", "Chemicals - Direct", "Fuel - Direct",
        "Lubricants - Direct", "Tools - Direct", "Dies / Moulds - Direct",
        "Subcontracting / Job Work - Direct", "Royalties - Direct",
        "Design / Drawing - Direct", "Special Material",
        "Other Direct Material", "Custom / Other",
    ],
    "Direct Labour": [
        "Direct Wages", "Production Salaries", "Machine Operators",
        "Assembly Labour", "Skilled Labour", "Semi-skilled Labour",
        "Unskilled Labour", "Contract Labour - Direct", "Overtime Wages",
        "Production Incentives", "Piece-rate Wages", "Shift Allowance",
        "Bonus - Direct Labour", "Employee Benefits - Direct Labour",
        "Other Direct Labour", "Custom / Other",
    ],
    "Direct Expenses": [
        "Special Machine Hire", "Special Equipment Hire",
        "Special Design Charges", "Testing Charges - Direct",
        "Inspection Charges - Direct", "Pattern / Tooling Charges",
        "Special Licence / Royalty", "Job Work Charges - Direct",
        "Freight Inward - Direct", "Other Direct Expense", "Custom / Other",
    ],
    "Factory Overhead": [
        "Indirect Materials", "Indirect Stores", "Indirect Consumables",
        "Indirect Tools", "Indirect Wages", "Factory Salaries",
        "Supervisors' Salaries", "Quality Control Salaries",
        "Factory Overtime", "Factory Employee Benefits", "Power", "Electricity",
        "Fuel", "Gas", "Steam", "Water", "Diesel", "Generator Expenses",
        "Machine Running Expenses", "Repairs & Maintenance - Machinery",
        "Repairs & Maintenance - Building", "Repairs & Maintenance - Electrical",
        "Factory Cleaning", "Factory Security", "Factory Insurance",
        "Factory Rent", "Factory Rates & Taxes", "Factory Depreciation",
        "Plant & Machinery Depreciation", "Factory Building Depreciation",
        "Factory Canteen", "Factory Safety Expenses", "Factory PPE",
        "Factory Calibration", "Factory Testing", "Factory Laboratory",
        "Factory IT / Software", "Factory Telephone / Internet",
        "Material Handling", "Internal Transport", "Forklift Expenses",
        "Factory Consumables", "Factory Scrap Handling", "Waste Disposal",
        "Production Planning", "Other Factory Overhead", "Custom / Other",
    ],
    "Administration Overhead": [
        "Office Salaries", "Management Salaries", "HR Salaries",
        "Finance & Accounts Salaries", "Administration Wages",
        "Employee Benefits - Administration", "Office Rent",
        "Office Electricity", "Office Water", "Office Repairs & Maintenance",
        "Office Insurance", "Office Depreciation", "Furniture Depreciation",
        "Computer Depreciation", "Printing & Stationery", "Postage & Courier",
        "Telephone", "Internet", "Software & Subscriptions", "IT Support",
        "Professional Fees", "Audit Fees", "Legal Fees", "Consultancy Fees",
        "Accounting Fees", "Bank Charges", "Office Travel", "Staff Welfare",
        "Training & Development", "Recruitment Expenses", "Office Security",
        "Office Cleaning", "Rates & Taxes - Administration",
        "Licences & Registration", "Corporate Expenses",
        "Other Administration Overhead", "Custom / Other",
    ],
    "Selling & Distribution": [
        "Sales Salaries", "Sales Commission", "Sales Incentives",
        "Marketing Salaries", "Advertising", "Digital Marketing",
        "Promotion Expenses", "Sales Promotion", "Exhibition & Events",
        "Samples", "Catalogue / Brochure", "Market Research",
        "Customer Relationship Expenses", "Freight Outward",
        "Transportation Outward", "Delivery Charges", "Loading & Unloading",
        "Warehousing", "Warehouse Rent", "Warehouse Salaries",
        "Warehouse Electricity", "Warehouse Insurance", "Packing - Selling",
        "Distribution Expenses", "Dealer Commission", "Agent Commission",
        "Bad Debts", "Sales Returns Handling", "After-sales Service",
        "Warranty Expenses", "Customer Support", "Sales Travel",
        "Sales Office Expenses", "Other Selling Expense",
        "Other Distribution Expense", "Custom / Other",
    ],
}

COST_LIBRARY["Cost Sheet Adjustments"] = [
    "Opening Work-in-Progress (WIP)",
    "Closing Work-in-Progress (WIP)",
    "Opening Finished Goods",
    "Closing Finished Goods",
]
CATEGORIES = list(COST_LIBRARY.keys())
CAT_SHORT = {
    "Direct Material": "Direct Material",
    "Direct Labour": "Direct Labour",
    "Direct Expenses": "Direct Expenses",
    "Factory Overhead": "Factory Overhead",
    "Administration Overhead": "Administration Overhead",
    "Selling & Distribution": "Selling & Distribution",
}

# =========================
# STYLING
# =========================
st.markdown("""
<style>
    .block-container {padding-top: 1.25rem; padding-bottom: 2rem; max-width: 1500px;}
    .hero {
        padding: 28px 32px; border-radius: 20px; margin-bottom: 18px;
        background: linear-gradient(135deg, #102A43 0%, #1F5A85 100%);
        color: white; box-shadow: 0 8px 28px rgba(16,42,67,.14);
    }
    .hero h1 {font-size: 2.25rem; margin: 0; font-weight: 800;}
    .hero p {margin: 7px 0 0; color: #D9EAF7; font-size: 1rem;}
    .section-title {
        font-size: 1.2rem; font-weight: 750; color: #102A43;
        margin: 18px 0 8px;
    }
    .info-box {
        border: 1px solid #D9E2EC; border-radius: 14px; padding: 12px 15px;
        background: #F7FAFC; margin-bottom: 10px;
    }
    .small-note {color:#52606D; font-size:.86rem;}
    div[data-testid="stMetric"] {
        border: 1px solid #D9E2EC; border-radius: 14px; padding: 14px;
        background: white; min-height: 92px; overflow: hidden;
    }
    div[data-testid="stMetricLabel"] {
        white-space: normal !important;
    }
    div[data-testid="stMetricValue"] {
        font-size: clamp(1.35rem, 2.2vw, 2rem) !important;
        white-space: nowrap !important; overflow: hidden !important; text-overflow: ellipsis !important;
    }
    button[kind="primary"] {font-weight: 700;}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hero">
  <h1>📊 Cost Sheet Pro</h1>
  <p>Professional, interactive standard cost-sheet generator • component-level selection • automated costing • Excel export</p>
</div>
""", unsafe_allow_html=True)

# =========================
# SESSION STATE
# =========================
if "cost_lines" not in st.session_state:
    st.session_state.cost_lines = [
        {"Category": "Direct Labour", "Cost Component": "Direct Wages", "Qty / Units": 1000.0, "Total Amount": 12000.0},
        {"Category": "Factory Overhead", "Cost Component": "Power", "Qty / Units": 1.0, "Total Amount": 8000.0},
    ]

# Backward compatibility: convert any older Qty/Rate rows to the new input model.
for _line in st.session_state.cost_lines:
    if "Total Amount" not in _line:
        qty_old = float(_line.get("Qty / Units", 0.0) or 0.0)
        rate_old = float(_line.get("Rate", 0.0) or 0.0)
        _line["Total Amount"] = qty_old * rate_old
    _line.pop("Rate", None)

# =========================
# SIDEBAR: PROJECT INPUTS
# =========================
with st.sidebar:
    st.header("⚙️ Cost Sheet Setup")
    company = st.text_input("Company / Organisation", "ABC Industries Ltd.")
    product = st.text_input("Product / Service", "Product / Service")
    period = st.text_input("Costing Period", "FY 2026-27")
    currency = st.text_input("Currency Symbol", "₹", max_chars=5)
    st.divider()

    st.subheader("Production & Inventory")
    units = st.number_input("Units Produced", min_value=0.0, value=1000.0, step=1.0)

    st.markdown("**Material Procurement**")
    purchases_rm = st.number_input("Raw Material Purchases", min_value=0.0, value=0.0, step=100.0)
    st.markdown("**Raw Material Stock**")
    opening_rm = st.number_input("Opening Raw Material Stock", min_value=0.0, value=0.0, step=100.0)
    closing_rm = st.number_input("Closing Raw Material Stock", min_value=0.0, value=0.0, step=100.0)

    st.markdown("**Work-in-Progress (WIP)**")
    opening_wip = st.number_input("Opening WIP", min_value=0.0, value=0.0, step=100.0)
    closing_wip = st.number_input("Closing WIP", min_value=0.0, value=0.0, step=100.0)

    st.markdown("**Finished Goods**")
    opening_fg = st.number_input("Opening Finished Goods", min_value=0.0, value=0.0, step=100.0)
    closing_fg = st.number_input("Closing Finished Goods", min_value=0.0, value=0.0, step=100.0)


# =========================
# ADD COMPONENT
# =========================
st.markdown('<div class="section-title">1. Add Individual Cost Components</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="info-box"><b>Select every cost individually.</b> Choose a category and cost component, enter the <b>Total Quantity / Units</b> and <b>Total Amount</b>. The system automatically calculates and displays the <b>Rate / Unit</b>.'
    'You can add the same component multiple times and use <b>Custom / Other</b> for anything not in the library.</div>',
    unsafe_allow_html=True
)

add_cols = st.columns([1.10, 2.00, 0.90, 1.20, 1.05, 1.00], gap="medium")
with add_cols[0]:
    add_category = st.selectbox("Category", CATEGORIES, key="add_category")
with add_cols[1]:
    add_component = st.selectbox(
        "Cost Component",
        COST_LIBRARY[add_category],
        key="add_component"
    )
with add_cols[2]:
    add_qty = st.number_input("Total Qty / Units", min_value=0.0, value=1.0, step=1.0, key="add_qty")
with add_cols[3]:
    add_total = st.number_input("Total Amount", min_value=0.0, value=0.0, step=100.0, key="add_total")
with add_cols[4]:
    add_auto_rate = add_total / add_qty if add_qty else 0.0
    st.number_input("Rate / Unit", min_value=0.0, value=float(add_auto_rate), disabled=True, key="add_auto_rate_display", format="%.2f")
with add_cols[5]:
    st.write("")
    st.write("")
    if st.button("➕ Add Cost", type="primary", use_container_width=True):
        st.session_state.cost_lines.append({
            "Category": add_category,
            "Cost Component": add_component,
            "Qty / Units": float(add_qty),
            "Total Amount": float(add_total),
        })
        st.rerun()

# Custom component entry
with st.expander("✏️ Add a Custom Cost Component", expanded=False):
    cc1, cc2, cc3, cc4, cc5 = st.columns([1.10, 2.00, 0.90, 1.20, 1.05], gap="medium")
    with cc1:
        custom_category = st.selectbox("Category", CATEGORIES, key="custom_category")
    with cc2:
        custom_name = st.text_input("Custom component name", placeholder="e.g., Special Compliance Fee")
    with cc3:
        custom_qty = st.number_input("Total Qty / Units", min_value=0.0, value=1.0, step=1.0, key="custom_qty")
    with cc4:
        custom_total = st.number_input("Total Amount", min_value=0.0, value=0.0, step=100.0, key="custom_total")
    with cc5:
        custom_auto_rate = custom_total / custom_qty if custom_qty else 0.0
        st.number_input("Rate / Unit", min_value=0.0, value=float(custom_auto_rate), disabled=True, key="custom_auto_rate_display", format="%.2f")
    if st.button("Add Custom Cost", use_container_width=True):
        if custom_name.strip():
            st.session_state.cost_lines.append({
                "Category": custom_category,
                "Cost Component": custom_name.strip(),
                "Qty / Units": float(custom_qty),
                "Total Amount": float(custom_total),
            })
            st.success(f"Added {custom_name.strip()}")
            st.rerun()
        else:
            st.error("Please enter a custom component name.")

# =========================
# COST LINES
# =========================
st.markdown('<div class="section-title">2. Cost Components & Values</div>', unsafe_allow_html=True)

if not st.session_state.cost_lines:
    st.info("No cost components yet. Use the selector above to add your first cost.")
else:
    display_df = pd.DataFrame(st.session_state.cost_lines)
    display_df["Qty / Units"] = pd.to_numeric(display_df.get("Qty / Units"), errors="coerce").fillna(0.0)
    display_df["Total Amount"] = pd.to_numeric(display_df.get("Total Amount"), errors="coerce").fillna(0.0)
    display_df["Rate / Unit"] = display_df.apply(
        lambda row: row["Total Amount"] / row["Qty / Units"] if row["Qty / Units"] else 0.0, axis=1
    )
    st.dataframe(
        display_df[["Category", "Cost Component", "Qty / Units", "Total Amount", "Rate / Unit"]].style.format({
            "Qty / Units": "{:,.2f}",
            "Total Amount": f"{currency}{{:,.2f}}",
            "Rate / Unit": f"{currency}{{:,.2f}}",
        }),
        use_container_width=True,
        hide_index=True,
        height=min(420, 90 + len(display_df) * 38),
    )

    remove_options = [
        f"{i+1}. {row['Category']} — {row['Cost Component']} — {currency}{float(row.get('Total Amount', 0.0)):,.2f}"
        for i, row in enumerate(st.session_state.cost_lines)
    ]
    r1, r2, r3 = st.columns([4.2, 1.3, 1.3])
    with r1:
        remove_choice = st.selectbox("Select a line to remove", ["— Select —"] + remove_options)
    with r2:
        st.write("")
        st.write("")
        if st.button("🗑️ Remove Selected", use_container_width=True):
            if remove_choice != "— Select —":
                idx = remove_options.index(remove_choice)
                st.session_state.cost_lines.pop(idx)
                st.rerun()
    with r3:
        st.write("")
        st.write("")
        if st.button("🧹 Clear All", use_container_width=True):
            st.session_state.cost_lines = []
            st.rerun()

st.markdown('<div class="section-title">3. Profit & Pricing</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="info-box"><b>Choose your profit method and margin.</b> The selected percentage is used directly in the sales-value calculation and is shown in the Excel output.</div>',
    unsafe_allow_html=True
)

profit_col1, profit_col2 = st.columns([1.55, 1.0], gap="large")
with profit_col1:
    profit_basis = st.radio(
        "Profit Margin Basis",
        ["Profit % on Cost", "Profit % on Sales"],
        horizontal=True,
        index=1,
        key="profit_basis_main",
        help="Select whether the profit percentage is based on Cost of Sales or on Sales Value."
    )
with profit_col2:
    profit_pct = st.number_input(
        "Profit Margin (%)",
        min_value=0.0,
        max_value=99.99,
        value=20.0,
        step=0.5,
        format="%.2f",
        key="profit_pct_main",
        help="Enter the margin you want. Example: 20 means 20%."
    )

if profit_basis == "Profit % on Sales":
    st.info(f"Selected: **{profit_pct:.2f}% profit on Sales** — Sales Value = Cost of Sales ÷ (1 − {profit_pct:.2f}%).")
else:
    st.info(f"Selected: **{profit_pct:.2f}% profit on Cost of Sales** — Profit = Cost of Sales × {profit_pct:.2f}%.")


# =========================
# CALCULATIONS
# =========================
df = pd.DataFrame(st.session_state.cost_lines)
if df.empty:
    df = pd.DataFrame(columns=["Category", "Cost Component", "Qty / Units", "Total Amount"])
df["Qty / Units"] = pd.to_numeric(df.get("Qty / Units", pd.Series(dtype=float)), errors="coerce").fillna(0.0)
df["Total Amount"] = pd.to_numeric(df.get("Total Amount", pd.Series(dtype=float)), errors="coerce").fillna(0.0)
df["Rate / Unit"] = df.apply(
    lambda row: row["Total Amount"] / row["Qty / Units"] if row["Qty / Units"] else 0.0, axis=1
)
df["Amount"] = df["Total Amount"]

def cat_total(category):
    return float(df.loc[df["Category"] == category, "Amount"].sum())

def selected_component_value(category, component, default=0.0):
    rows = df.loc[
        (df["Category"] == category) &
        (df["Cost Component"] == component),
        "Amount"
    ]
    return float(rows.sum()) if not rows.empty else float(default)

# Direct Material stock/purchase inputs are selectable under Direct Material.
# Sidebar values are fallback values when the user has not added those lines.
opening_rm_calc = selected_component_value("Direct Material", "Opening Raw Material Stock", opening_rm)
purchases_rm_calc = selected_component_value("Direct Material", "Raw Material Purchases", purchases_rm)
closing_rm_calc = selected_component_value("Direct Material", "Closing Raw Material Stock", closing_rm)

# WIP and finished-goods adjustments are selectable under Cost Sheet Adjustments.
opening_wip_calc = selected_component_value("Cost Sheet Adjustments", "Opening Work-in-Progress (WIP)", opening_wip)
closing_wip_calc = selected_component_value("Cost Sheet Adjustments", "Closing Work-in-Progress (WIP)", closing_wip)
opening_fg_calc = selected_component_value("Cost Sheet Adjustments", "Opening Finished Goods", opening_fg)
closing_fg_calc = selected_component_value("Cost Sheet Adjustments", "Closing Finished Goods", closing_fg)

# Other direct-material components exclude the three inventory/purchase lines above.
direct_material_other = float(
    df.loc[
        (df["Category"] == "Direct Material") &
        (~df["Cost Component"].isin([
            "Opening Raw Material Stock",
            "Raw Material Purchases",
            "Closing Raw Material Stock"
        ])),
        "Amount"
    ].sum()
)

material_consumed = opening_rm_calc + purchases_rm_calc - closing_rm_calc
material = material_consumed + direct_material_other

labour = cat_total("Direct Labour")
direct_exp = cat_total("Direct Expenses")
factory_oh = cat_total("Factory Overhead")
admin_oh = cat_total("Administration Overhead")
selling_oh = cat_total("Selling & Distribution")

prime_cost = material + labour + direct_exp
works_cost = prime_cost + factory_oh + opening_wip_calc - closing_wip_calc
cost_of_production = works_cost + admin_oh
cost_of_goods_sold = cost_of_production + opening_fg_calc - closing_fg_calc
cost_of_sales = cost_of_goods_sold + selling_oh
if profit_basis == "Profit % on Cost":
    profit = cost_of_sales * profit_pct / 100.0
    sales_value_before_tax = cost_of_sales + profit
else:
    sales_value_before_tax = cost_of_sales / (1.0 - profit_pct / 100.0)
    profit = sales_value_before_tax - cost_of_sales

# GST/Tax is intentionally not included in the final cost-sheet Excel output.
# It is not used in the cost-sheet totals to avoid presenting tax as a cost.
tax = 0.0
final_invoice_value = sales_value_before_tax

# =========================
# DASHBOARD
# =========================
st.markdown('<div class="section-title">4. Live Cost Dashboard</div>', unsafe_allow_html=True)
m1 = st.columns(3, gap="large")
for col, (label, value) in zip(m1, [
    ("Prime Cost", prime_cost),
    ("Works Cost", works_cost),
    ("Cost of Production", cost_of_production),
]):
    col.metric(label, f"{currency}{value:,.2f}")

m2 = st.columns(3, gap="large")
for col, (label, value) in zip(m2, [
    ("Cost of Sales", cost_of_sales),
    ("Profit", profit),
    ("Sales Value", sales_value_before_tax),
]):
    col.metric(label, f"{currency}{value:,.2f}")

u = st.columns(4, gap="large")
unit_cost = cost_of_sales / units if units else 0
unit_price = sales_value_before_tax / units if units else 0
unit_profit = profit / units if units else 0
u[0].metric("Cost / Unit", f"{currency}{unit_cost:,.2f}" if units else "—")
u[1].metric("Selling Price / Unit", f"{currency}{unit_price:,.2f}" if units else "—")
u[2].metric("Profit / Unit", f"{currency}{unit_profit:,.2f}" if units else "—")
u[3].metric("Sales Value", f"{currency}{sales_value_before_tax:,.2f}")

# =========================
# STANDARD COST SHEET
# =========================
st.markdown('<div class="section-title">5. Standard Cost Sheet</div>', unsafe_allow_html=True)

summary_rows = [
    ("OPENING RAW MATERIAL STOCK", opening_rm_calc, "adjustment"),
    ("ADD: RAW MATERIAL PURCHASES", purchases_rm_calc, "adjustment"),
    ("LESS: CLOSING RAW MATERIAL STOCK", -closing_rm_calc, "adjustment"),
    ("RAW MATERIAL CONSUMED", material_consumed, "total"),
    ("OTHER DIRECT MATERIAL", direct_material_other, "section"),
    ("DIRECT MATERIAL", material, "section"),
    ("DIRECT LABOUR", labour, "section"),
    ("DIRECT EXPENSES", direct_exp, "section"),
    ("PRIME COST", prime_cost, "total"),
    ("FACTORY / WORKS OVERHEAD", factory_oh, "section"),
    ("ADD: OPENING WIP", opening_wip_calc, "adjustment"),
    ("LESS: CLOSING WIP", -closing_wip_calc, "adjustment"),
    ("WORKS / FACTORY COST", works_cost, "total"),
    ("ADMINISTRATION OVERHEAD", admin_oh, "section"),
    ("COST OF PRODUCTION", cost_of_production, "total"),
    ("ADD: OPENING FINISHED GOODS", opening_fg_calc, "adjustment"),
    ("LESS: CLOSING FINISHED GOODS", -closing_fg_calc, "adjustment"),
    ("COST OF GOODS SOLD", cost_of_goods_sold, "total"),
    ("SELLING & DISTRIBUTION OVERHEAD", selling_oh, "section"),
    ("COST OF SALES", cost_of_sales, "total"),
    (f"ADD: PROFIT — {profit_pct:.2f}% {'ON COST' if profit_basis == 'Profit % on Cost' else 'ON SALES'}", profit, "profit"),
    ("SALES VALUE", sales_value_before_tax, "grand"),
]
summary_df = pd.DataFrame(summary_rows, columns=["Cost Head", "Amount", "Type"])
summary_df["% of Sales"] = summary_df["Amount"].apply(
    lambda x: x / sales_value_before_tax if sales_value_before_tax else 0
)

st.dataframe(
    summary_df[["Cost Head", "Amount", "% of Sales"]].style.format({
        "Amount": f"{currency}{{:,.2f}}",
        "% of Sales": "{:.1%}",
    }),
    use_container_width=True,
    hide_index=True,
    height=690,
)

# =========================
# COST ANALYSIS TABS
# =========================
tab1, tab2, tab3 = st.tabs(["📈 Category Analysis", "🔎 Component Detail", "ℹ️ Methodology"])

with tab1:
    cat_df = pd.DataFrame({
        "Category": ["Direct Material", "Direct Labour", "Direct Expenses",
                     "Factory Overhead", "Administration Overhead", "Selling & Distribution"],
        "Amount": [material, labour, direct_exp, factory_oh, admin_oh, selling_oh]
    })
    st.bar_chart(cat_df.set_index("Category"))
    st.dataframe(
        cat_df.style.format({"Amount": f"{currency}{{:,.2f}}"}),
        use_container_width=True, hide_index=True
    )

with tab2:
    if not df.empty:
        detail = df[["Category", "Cost Component", "Qty / Units", "Total Amount", "Rate / Unit", "Amount"]].copy()
        st.dataframe(
            detail.style.format({
                "Qty / Units": "{:,.2f}",
                "Total Amount": f"{currency}{{:,.2f}}",
                "Rate / Unit": f"{currency}{{:,.2f}}",
                "Amount": f"{currency}{{:,.2f}}",
            }),
            use_container_width=True, hide_index=True,
        )
    else:
        st.info("Add cost components to see component-level analysis.")

with tab3:
    st.markdown("""
**Costing flow used by this model**

**Component input method:** Total Amount is entered by the user; Rate / Unit is calculated automatically as Total Amount ÷ Total Qty / Units.

1. **Raw Material Consumed** = Opening Raw Material Stock + Raw Material Purchases − Closing Raw Material Stock
2. **Direct Material** = Raw Material Consumed + Other Direct Material Components
3. **Prime Cost** = Direct Material + Direct Labour + Direct Expenses
4. **Works / Factory Cost** = Prime Cost + Factory Overhead + Opening WIP − Closing WIP
5. **Cost of Production** = Works / Factory Cost + Administration Overhead
6. **Cost of Goods Sold** = Cost of Production + Opening Finished Goods − Closing Finished Goods
7. **Cost of Sales** = Cost of Goods Sold + Selling & Distribution Overhead
8. **Profit on Cost** = Cost of Sales × Profit % on Cost
9. **Profit on Sales** = Sales Value × Profit % on Sales
10. **Sales Value** = Cost of Sales + Profit (on cost) OR Cost of Sales ÷ (1 − Profit % on Sales)

**Tax note:** GST/Tax is deliberately excluded from this cost-sheet model and Excel output. It is a statutory billing/tax item, not a component of product cost. This prevents the cost sheet from misleadingly presenting GST as cost or profit.
""")

# =========================
# EXCEL EXPORT
# =========================
def build_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Sheet"
    detail_ws = wb.create_sheet("Cost Components")
    inputs_ws = wb.create_sheet("Inputs & Assumptions")

    # Palette
    NAVY = "102A43"
    BLUE = "1F5A85"
    TEAL = "0F6B78"
    LIGHT = "EAF2F8"
    PALE = "F7FAFC"
    GREEN = "E2F0D9"
    ORANGE = "FCE4D6"
    RED = "F4CCCC"
    WHITE = "FFFFFF"
    GREY = "52606D"
    BORDER = "B7C4CE"
    thin = Side(style="thin", color=BORDER)

    # Cost Sheet columns
    widths = [6, 48, 16, 18, 20, 16]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"

    ws.merge_cells("A1:F1")
    ws["A1"] = company
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"] = "STANDARD COST SHEET"
    ws["A2"].font = Font(size=15, bold=True, color=WHITE)
    ws["A2"].fill = PatternFill("solid", fgColor=BLUE)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:F3")
    ws["A3"] = f"{product}  |  {period}"
    ws["A3"].font = Font(size=11, italic=True, color=GREY)
    ws["A3"].alignment = Alignment(horizontal="center")

    headers = ["No.", "Cost Component", "Qty / Units", "Rate / Unit", "Amount", "% of Sales"]
    for c, header in enumerate(headers, 1):
        cell = ws.cell(5, c, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=thin, bottom=thin)

    r = 6
    line_no = 1
    for category in CATEGORIES:
        subset = df[df["Category"] == category] if not df.empty else pd.DataFrame()
        if subset.empty:
            continue
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        cell = ws.cell(r, 2, category.upper())
        cell.font = Font(bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=LIGHT)
        for c in range(1, 7):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=LIGHT)
        r += 1

        for _, item in subset.iterrows():
            ws.cell(r, 1, line_no)
            ws.cell(r, 2, str(item["Cost Component"]))
            ws.cell(r, 3, float(item["Qty / Units"]))
            ws.cell(r, 4, float(item["Rate / Unit"]))
            ws.cell(r, 5, float(item["Total Amount"]))
            # Formula refers to sales value in the summary section, filled later.
            for c in range(1, 7):
                ws.cell(r, c).border = Border(bottom=thin)
                ws.cell(r, c).alignment = Alignment(vertical="center")
            for c in [3, 4, 5]:
                ws.cell(r, c).number_format = '#,##0.00'
            line_no += 1
            r += 1

    r += 1
    summary_start = r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(r, 1, "STANDARD COST SUMMARY")
    ws.cell(r, 1).font = Font(size=13, bold=True, color=WHITE)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(r, 1).alignment = Alignment(horizontal="left")
    for c in range(1, 7):
        ws.cell(r, c).fill = PatternFill("solid", fgColor=NAVY)
    r += 1

    summary_excel_rows = {}
    for label, val, kind in summary_rows:
        summary_excel_rows[label] = r
        ws.cell(r, 2, label)
        ws.cell(r, 5, float(val))
        ws.cell(r, 5).number_format = '#,##0.00'
        # Percentage-of-sales formula is applied after the SALES VALUE row is known.
        ws.cell(r, 6).number_format = '0.0%'
        if kind == "section":
            fill = LIGHT
            font_color = NAVY
            bold = True
        elif kind == "total":
            fill = GREEN
            font_color = NAVY
            bold = True
        elif kind == "profit":
            fill = ORANGE
            font_color = NAVY
            bold = True
        else:
            fill = WHITE
            font_color = "000000"
            bold = False
        for c in range(1, 7):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=fill)
            ws.cell(r, c).border = Border(bottom=thin)
        ws.cell(r, 2).font = Font(bold=bold, color=font_color)
        ws.cell(r, 5).font = Font(bold=bold, color=font_color)
        r += 1

    sales_row = summary_excel_rows["SALES VALUE"]
    # Fix all percentage formulas now that sales row is known.
    for row_num in range(summary_start + 1, r):
        ws.cell(row_num, 6, f'=IFERROR(E{row_num}/E{sales_row},0)')
        ws.cell(row_num, 6).number_format = '0.0%'

    r += 2
    ws.cell(r, 2, "UNIT ECONOMICS")
    ws.cell(r, 2).font = Font(size=13, bold=True, color=NAVY)
    unit_rows = [
        ("Units Produced", units, False),
        ("Opening Raw Material Stock", opening_rm, False),
        ("Raw Material Purchases", purchases_rm, False),
        ("Closing Raw Material Stock", closing_rm, False),
        ("Opening WIP", opening_wip, False),
        ("Closing WIP", closing_wip, False),
        ("Opening Finished Goods", opening_fg, False),
        ("Closing Finished Goods", closing_fg, False),
        ("Cost / Unit", unit_cost, False),
        ("Sales Value / Unit", unit_price, False),
        ("Profit / Unit", unit_profit, False),
        ("Profit Basis", profit_basis, False),
        ("Profit Margin %", profit_pct, True),
    ]
    for label, val, is_pct in unit_rows:
        r += 1
        ws.cell(r, 2, label)
        if isinstance(val, (int, float)):
            ws.cell(r, 5, float(val))
            ws.cell(r, 5).number_format = '0.0%' if is_pct else '#,##0.00'
        else:
            ws.cell(r, 5, str(val))
            ws.cell(r, 5).number_format = '@'
        ws.cell(r, 2).font = Font(bold=True if "Unit" in label or "%" in label else False)

    # Cost components sheet
    detail_ws.sheet_view.showGridLines = False
    detail_ws.freeze_panes = "A2"
    detail_headers = ["Category", "Cost Component", "Qty / Units", "Rate / Unit", "Total Amount"]
    for c, header in enumerate(detail_headers, 1):
        cell = detail_ws.cell(1, c, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center")
    for i, width in enumerate([26, 42, 15, 18, 20], 1):
        detail_ws.column_dimensions[get_column_letter(i)].width = width
    for rr, (_, item) in enumerate(df.iterrows(), 2):
        values = [item["Category"], item["Cost Component"], float(item["Qty / Units"]),
                  float(item["Rate / Unit"]), float(item["Total Amount"])]
        for c, value in enumerate(values, 1):
            detail_ws.cell(rr, c, value)
            detail_ws.cell(rr, c).border = Border(bottom=thin)
        detail_ws.cell(rr, 3).number_format = '#,##0.00'
        detail_ws.cell(rr, 4).number_format = '#,##0.00'
        detail_ws.cell(rr, 5).number_format = '#,##0.00'

    # Inputs sheet
    inputs_ws.sheet_view.showGridLines = False
    inputs_ws.column_dimensions["A"].width = 34
    inputs_ws.column_dimensions["B"].width = 28
    inputs_ws["A1"] = "INPUTS & ASSUMPTIONS"
    inputs_ws["A1"].font = Font(size=16, bold=True, color=WHITE)
    inputs_ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    inputs_ws.merge_cells("A1:B1")
    input_rows = [
        ("Company / Organisation", company),
        ("Product / Service", product),
        ("Costing Period", period),
        ("Currency Symbol", currency),
        ("Units Produced", units),
        ("Opening Raw Material Stock", opening_rm),
        ("Raw Material Purchases", purchases_rm),
        ("Closing Raw Material Stock", closing_rm),
        ("Opening WIP", opening_wip),
        ("Closing WIP", closing_wip),
        ("Opening Finished Goods", opening_fg),
        ("Closing Finished Goods", closing_fg),
        ("Profit Basis", profit_basis),
        ("Profit Margin %", profit_pct / 100),
    ]
    for rr, (label, value) in enumerate(input_rows, 3):
        inputs_ws.cell(rr, 1, label)
        inputs_ws.cell(rr, 2, value)
        inputs_ws.cell(rr, 1).font = Font(bold=True, color=NAVY)
        inputs_ws.cell(rr, 1).fill = PatternFill("solid", fgColor=LIGHT)
        inputs_ws.cell(rr, 1).border = Border(bottom=thin)
        inputs_ws.cell(rr, 2).border = Border(bottom=thin)
        if "%" in label:
            inputs_ws.cell(rr, 2).number_format = '0.0%'
        elif isinstance(value, (int, float)):
            inputs_ws.cell(rr, 2).number_format = '#,##0.00'

    # Print setup
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_title_rows = "1:5"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

st.markdown('<div class="section-title">6. Export</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="info-box">Your Excel workbook contains a formatted <b>Cost Sheet</b>, '
    '<b>Cost Components</b> detail sheet, and <b>Inputs & Assumptions</b> sheet.</div>',
    unsafe_allow_html=True
)
excel_bytes = build_excel()
st.download_button(
    "⬇️ Download Professional Excel Cost Sheet",
    data=excel_bytes,
    file_name=f"{company.replace(' ', '_')}_Cost_Sheet.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
    use_container_width=True,
)
